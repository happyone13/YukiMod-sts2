import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CARD_DIR = ROOT / "YukiModCode" / "Cards"
LOC_PATH = ROOT / "YukiMod" / "localization" / "zhs" / "cards.json"
XLSX_PATH = ROOT / "docs" / "yuki-card-table.xlsx"
MD_PATH = ROOT / "docs" / "yuki-card-table.zh-CN.md"

RARITY_CN = {
    "Basic": "初始",
    "Common": "通常",
    "Uncommon": "罕见",
    "Rare": "稀有",
    "Ancient": "先古",
    "Token": "Token",
}
TYPE_CN = {"Attack": "攻击", "Skill": "技能", "Power": "能力"}
SCHOOL_CN = {
    "Inspiration": "灵感",
    "Moonshadow": "月影",
    "BlackCloud": "黑云",
    "Other": "其他",
}
POOL_CN = {"YukiModCardPool": "主牌池", "TokenCardPool": "Token牌池"}

RARITY_ORDER = ["初始", "通常", "罕见", "稀有", "先古", "Token"]
TYPE_ORDER = ["攻击", "技能", "能力"]
SCHOOL_ORDER = ["灵感", "月影", "黑云", "其他"]


def pascal_to_id(name: str) -> str:
    return re.sub(r"(?<!^)(?=[A-Z])", "_", name).upper()


def parse_number(text: str):
    text = text.strip().rstrip("mM")
    try:
        value = float(text)
    except ValueError:
        return None
    return int(value) if value.is_integer() else value


def number_text(value) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).rstrip("0").rstrip(".")


def split_args(arg_text: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    escaped = False
    for char in arg_text:
        if in_string:
            current.append(char)
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
            current.append(char)
        elif char in "(<[":
            depth += 1
            current.append(char)
        elif char in ")>]":
            depth -= 1
            current.append(char)
        elif char == "," and depth == 0:
            parts.append("".join(current).strip())
            current = []
        else:
            current.append(char)

    if current:
        parts.append("".join(current).strip())
    return parts


def parse_constants(text: str) -> dict[str, str]:
    return {
        name: value
        for name, value in re.findall(
            r'(?:private\s+|protected\s+|public\s+)?const\s+string\s+(\w+)\s*=\s*"([^"]+)"',
            text,
        )
    }


def resolve_name(expr: str, constants: dict[str, str]) -> str:
    expr = expr.strip()
    if expr.startswith('"') and expr.endswith('"'):
        return expr[1:-1]
    return constants.get(expr, expr)


def parse_vars(text: str) -> dict[str, dict[str, float]]:
    constants = parse_constants(text)
    variables: dict[str, dict[str, float]] = {}
    prop_to_name = {
        "Damage": "Damage",
        "Block": "Block",
        "Cards": "Cards",
        "Energy": "Energy",
        "Repeat": "Repeat",
        "Weak": "WeakPower",
        "Strength": "StrengthPower",
        "Vulnerable": "VulnerablePower",
    }

    for var_type, args in re.findall(
        r"new\s+(DamageVar|BlockVar|CardsVar|EnergyVar|RepeatVar)\s*\(([^\)]*)\)",
        text,
        re.S,
    ):
        value = parse_number(split_args(args)[0])
        if value is not None:
            variables[var_type[:-3]] = {"base": value, "delta": 0}

    for power, args in re.findall(r"new\s+PowerVar<([A-Za-z0-9_]+)>\s*\(([^\)]*)\)", text, re.S):
        value = parse_number(split_args(args)[0])
        if value is not None:
            variables[power] = {"base": value, "delta": 0}

    for _, args in re.findall(r"new\s+(DynamicVar|IntVar)\s*\(([^\)]*)\)", text, re.S):
        parts = split_args(args)
        if len(parts) < 2:
            continue
        name = resolve_name(parts[0], constants)
        value = parse_number(parts[1])
        if value is not None:
            variables[name] = {"base": value, "delta": 0}

    for prop, delta in re.findall(
        r"DynamicVars\.([A-Za-z0-9_]+)\.UpgradeValueBy\((-?\d+(?:\.\d+)?)m?\)",
        text,
    ):
        name = prop_to_name.get(prop, prop)
        if name in variables:
            variables[name]["delta"] += parse_number(delta) or 0

    for key, delta in re.findall(
        r'DynamicVars\["([^"]+)"\]\.UpgradeValueBy\((-?\d+(?:\.\d+)?)m?\)',
        text,
    ):
        if key in variables:
            variables[key]["delta"] += parse_number(delta) or 0

    for key_expr, delta in re.findall(
        r"DynamicVars\[(\w+)\]\.UpgradeValueBy\((-?\d+(?:\.\d+)?)m?\)",
        text,
    ):
        key = constants.get(key_expr, key_expr)
        if key in variables:
            variables[key]["delta"] += parse_number(delta) or 0

    return variables


def clean_description(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "；")
    text = re.sub(r"\[(?:/)?[a-zA-Z_]+\]", "", text)
    while "；；" in text:
        text = text.replace("；；", "；")
    return text.strip("； ")


def substitute_dynamic_text(description: str, variables: dict[str, dict[str, float]]) -> str:
    def repl(match: re.Match[str]) -> str:
        raw = match.group(1)
        name = raw.split(":", 1)[0]
        if name not in variables:
            return "{" + raw + "}"

        base = variables[name]["base"]
        upgraded = base + variables[name]["delta"]
        if variables[name]["delta"]:
            return f"{number_text(base)}/{number_text(upgraded)}"
        return number_text(base)

    return re.sub(r"\{([^{}]+)\}", repl, description)


def parse_keywords(text: str) -> list[str]:
    keywords: list[str] = []
    if "CanonicalKeywords" in text:
        for keyword in re.findall(r"CardKeyword\.([A-Za-z0-9_]+)", text):
            if keyword not in keywords:
                keywords.append(keyword)
    for keyword in re.findall(r"AddKeyword\(CardKeyword\.([A-Za-z0-9_]+)\)", text):
        if keyword not in keywords:
            keywords.append(keyword)
    return keywords


def load_cards() -> list[dict[str, object]]:
    loc = json.loads(LOC_PATH.read_text(encoding="utf-8"))
    cards: list[dict[str, object]] = []

    for path in sorted(CARD_DIR.glob("*.cs"), key=lambda item: item.name.lower()):
        text = path.read_text(encoding="utf-8")
        pool_match = re.search(r"\[Pool\(typeof\(([^)]+)\)\)\]", text)
        ctor = re.search(
            r"public\s+class\s+(\w+)\s*\(\)\s*:\s*(YukiModTokenCard|YukiModCard)"
            r"\s*\(\s*([^,]+)\s*,\s*CardType\.(\w+)\s*,\s*CardRarity\.(\w+)",
            text,
        )
        if not pool_match or not ctor:
            continue

        pool = pool_match.group(1)
        if pool not in POOL_CN:
            continue

        cls = ctor.group(1)
        card_id = pascal_to_id(cls)
        title = loc.get(f"YUKIMOD-{card_id}.title", cls)
        raw_description = loc.get(f"YUKIMOD-{card_id}.description", "")
        school_match = re.search(r"YukiCardSchool\.([A-Za-z0-9_]+)", text)
        school = SCHOOL_CN.get(school_match.group(1) if school_match else "Other", "其他")

        effect = substitute_dynamic_text(clean_description(raw_description), parse_vars(text))
        notes = [f"ID: YUKIMOD-{card_id}", f"类: {cls}", f"牌池: {POOL_CN[pool]}"]

        keywords = parse_keywords(text)
        if keywords:
            notes.append("源码关键词: " + ", ".join(keywords))
        cost_upgrade = re.search(r"EnergyCost\.UpgradeBy\((-?\d+)\)", text)
        if cost_upgrade:
            notes.append(f"升级费用变化: {cost_upgrade.group(1)}")
        if "RemoveKeyword(CardKeyword.Exhaust)" in text:
            notes.append("升级移除消耗")

        cost = parse_number(ctor.group(3))
        cards.append(
            {
                "title": title,
                "cost": cost if cost is not None else ctor.group(3).strip(),
                "rarity": RARITY_CN.get(ctor.group(5), ctor.group(5)),
                "type": TYPE_CN.get(ctor.group(4), ctor.group(4)),
                "effect": effect,
                "school": school,
                "notes": "；".join(notes),
                "score": None,
                "class": cls,
                "pool": pool,
            }
        )

    cards.sort(
        key=lambda card: (
            0 if card["pool"] == "YukiModCardPool" else 1,
            RARITY_ORDER.index(card["rarity"]) if card["rarity"] in RARITY_ORDER else 99,
            SCHOOL_ORDER.index(card["school"]) if card["school"] in SCHOOL_ORDER else 99,
            TYPE_ORDER.index(card["type"]) if card["type"] in TYPE_ORDER else 99,
            str(card["title"]),
        )
    )
    return cards


def style_sheet(ws, widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D6DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(cards: list[dict[str, object]], notes: list[tuple[str, str]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("卡牌表")
    headers = ["名称", "费用", "稀有度", "种类", "效果", "流派", "备注", "评分"]
    ws.append(headers)
    for card in cards:
        ws.append(
            [
                card["title"],
                card["cost"],
                card["rarity"],
                card["type"],
                card["effect"],
                card["school"],
                card["notes"],
                card["score"],
            ]
        )
    style_sheet(ws, [18, 8, 10, 8, 72, 10, 52, 10])
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 48

    ws_school = wb.create_sheet("流派统计")
    ws_school.append(["流派", "数量"] + RARITY_ORDER + TYPE_ORDER)
    for school in SCHOOL_ORDER:
        group = [card for card in cards if card["school"] == school]
        ws_school.append(
            [school, len(group)]
            + [sum(card["rarity"] == rarity for card in group) for rarity in RARITY_ORDER]
            + [sum(card["type"] == card_type for card in group) for card_type in TYPE_ORDER]
        )
    style_sheet(ws_school, [12, 10] + [10] * (len(RARITY_ORDER) + len(TYPE_ORDER)))

    ws_rarity = wb.create_sheet("稀有度统计")
    ws_rarity.append(["稀有度", "数量"] + SCHOOL_ORDER + TYPE_ORDER)
    for rarity in RARITY_ORDER:
        group = [card for card in cards if card["rarity"] == rarity]
        ws_rarity.append(
            [rarity, len(group)]
            + [sum(card["school"] == school for card in group) for school in SCHOOL_ORDER]
            + [sum(card["type"] == card_type for card in group) for card_type in TYPE_ORDER]
        )
    style_sheet(ws_rarity, [12, 10] + [10] * (len(SCHOOL_ORDER) + len(TYPE_ORDER)))

    ws_type = wb.create_sheet("种类统计")
    ws_type.append(["种类", "数量"] + RARITY_ORDER + SCHOOL_ORDER)
    for card_type in TYPE_ORDER:
        group = [card for card in cards if card["type"] == card_type]
        ws_type.append(
            [card_type, len(group)]
            + [sum(card["rarity"] == rarity for card in group) for rarity in RARITY_ORDER]
            + [sum(card["school"] == school for card in group) for school in SCHOOL_ORDER]
        )
    style_sheet(ws_type, [12, 10] + [10] * (len(RARITY_ORDER) + len(SCHOOL_ORDER)))

    ws_notes = wb.create_sheet("说明")
    ws_notes.append(["项目", "内容"])
    for note in notes:
        ws_notes.append(list(note))
    style_sheet(ws_notes, [18, 100])
    for row in range(2, ws_notes.max_row + 1):
        ws_notes.row_dimensions[row].height = 36

    ws_key = wb.create_sheet("关键词，能力与遗物")
    ws_key.append(["项目", "内容", "备注"])
    for row in [
        ("灵感", "回合开始以外加入手牌时会激活卡牌的灵感效果。", "关键词/机制"),
        ("黑云", "处于黑云姿态时触发黑云效果；使用非攻击牌会退出黑云姿态。", "关键词/机制"),
        ("凝聚", "若手中没有月影则生成月影；否则强化手中的月影。", "关键词/机制"),
        ("预见", "查看抽牌堆顶部若干牌，并可将其中任意张置入弃牌堆。", "关键词/机制"),
    ]:
        ws_key.append(list(row))
    style_sheet(ws_key, [18, 80, 18])

    wb.save(XLSX_PATH)
    load_workbook(XLSX_PATH, read_only=True).close()


def md_escape(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", "<br>").replace("|", "\\|")


def write_markdown(cards: list[dict[str, object]], notes: list[tuple[str, str]]) -> None:
    by_school = Counter(card["school"] for card in cards)
    by_rarity = Counter(card["rarity"] for card in cards)
    by_type = Counter(card["type"] for card in cards)

    lines: list[str] = [
        "# Yuki 卡牌实现表",
        "",
        "说明：",
    ]
    lines.extend(f"- {content}" for _, content in notes)
    lines.extend(
        [
            "",
            "## 卡牌表",
            "",
            "| 名称 | 费用 | 稀有度 | 种类 | 效果 | 流派 | 备注 | 评分 |",
            "| --- | ---: | --- | --- | --- | --- | --- | ---: |",
        ]
    )
    for card in cards:
        values = [
            card["title"],
            card["cost"],
            card["rarity"],
            card["type"],
            card["effect"],
            card["school"],
            card["notes"],
            card["score"],
        ]
        lines.append("| " + " | ".join(md_escape(value) for value in values) + " |")

    lines.extend(["", "## 统计", "", "### 流派", "", "| 流派 | 数量 |", "| --- | ---: |"])
    lines.extend(f"| {school} | {by_school[school]} |" for school in SCHOOL_ORDER)

    lines.extend(["", "### 稀有度", "", "| 稀有度 | 数量 |", "| --- | ---: |"])
    lines.extend(f"| {rarity} | {by_rarity[rarity]} |" for rarity in RARITY_ORDER)

    lines.extend(["", "### 种类", "", "| 种类 | 数量 |", "| --- | ---: |"])
    lines.extend(f"| {card_type} | {by_type[card_type]} |" for card_type in TYPE_ORDER)

    MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    cards = load_cards()
    notes = [
        ("生成口径", "卡牌行由 YukiModCode/Cards 中挂到 YukiModCardPool 或 TokenCardPool 的卡牌类生成。"),
        ("费用/稀有度/种类", "来自卡牌构造函数。费用记录基础费用数字。"),
        ("流派", "来自卡牌类的 YukiCardSchool；未重写则记为“其他”。"),
        (
            "效果",
            "来自 YukiMod/localization/zhs/cards.json，并按源码 DynamicVars 尽量替换基础/升级数值；未能静态展开的运行时变量保留 {变量名}。",
        ),
        ("Token", "Token 牌也列入本表，但牌池标记为 Token牌池。"),
        ("评分", "评分不是游戏运行内容，本次不从旧表沿用，避免错误分数伪装成当前实现。"),
    ]
    write_workbook(cards, notes)
    write_markdown(cards, notes)

    print(f"cards={len(cards)}")
    print("rarity=" + ", ".join(f"{rarity}:{sum(card['rarity'] == rarity for card in cards)}" for rarity in RARITY_ORDER))
    print("type=" + ", ".join(f"{card_type}:{sum(card['type'] == card_type for card in cards)}" for card_type in TYPE_ORDER))
    print("school=" + ", ".join(f"{school}:{sum(card['school'] == school for card in cards)}" for school in SCHOOL_ORDER))


if __name__ == "__main__":
    main()
